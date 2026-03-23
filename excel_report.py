import logging
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

COLOR = {
    "header_bg":  "1F3864",
    "ti_blue":    "2E75B6",
    "drop_in":    "C6EFCE",
    "p2p":        "DDEBF7",
    "similar":    "FFF2CC",
    "same_fn":    "FCE4D6",
    "no_match":   "F2F2F2",
    "row_alt":    "F9F9F9",
    "row_white":  "FFFFFF",
    "border":     "D9D9D9",
    "green_txt":  "375623",
}

MATCH_COLOR = {
    "Drop-in Replacement": "C6EFCE",
    "Pin-to-Pin":          "DDEBF7",
    "Similar":             "FFF2CC",
    "Same Functionality":  "FCE4D6",
}

SUMMARY_COLS = [
    ("EOL Part Number",        22),
    ("Supplier",               20),
    ("Description",            36),
    ("EOL Date",               14),
    ("Last Buy Date",          14),
    ("Supplier Replacement",   22),
    ("TI Alternative",         22),
    ("TI Description",         36),
    ("Match Type",             22),
    ("Lifecycle",              14),
    ("Vin Min",                12),
    ("Vin Max",                12),
    ("Vout Min",               12),
    ("Vout Max",               12),
    ("Iout",                   12),
    ("Package",                14),
    ("TI Product Page",        42),
    ("EOL Source / Reference", 50),
]


def _border():
    s = Side(style="thin", color=COLOR["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _hcell(ws, row, col, val, bg):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _border()
    return c


def _dcell(ws, row, col, val, bg="FFFFFF", bold=False, color="000000", wrap=False, link=None):
    c = ws.cell(row=row, column=col, value=val)
    if link:
        c.hyperlink = link
        c.font = Font(color="0563C1", underline="single", size=10, name="Arial")
    else:
        c.font = Font(bold=bold, color=color, size=10, name="Arial")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="center", wrap_text=wrap)
    c.border = _border()
    return c


def _fmt_date(d):
    if not d:
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%m/%d/%Y")
    return str(d)


def _source(eol):
    url = (eol.source_url or "").strip()
    doc = (eol.source_doc or "").strip()
    if url:
        parsed = urlparse(url)
        domain = parsed.netloc.replace("www.", "")
        tail = parsed.path.rstrip("/").split("/")[-1] or ""
        label = f"{domain} — {tail}" if tail else domain
        return label, url
    if doc:
        return doc, ""
    return "—", ""


def _row_bg(match_type):
    return MATCH_COLOR.get(match_type, COLOR["no_match"])


def generate(results: list, output_path: str) -> str:
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Summary"
    ws.freeze_panes = "A3"

    ws.merge_cells(f"A1:{get_column_letter(len(SUMMARY_COLS))}1")
    tc = ws.cell(row=1, column=1,
                 value=f"EOL → TI Cross-Reference Report  |  {datetime.now().strftime('%B %d, %Y')}")
    tc.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    tc.fill = PatternFill("solid", fgColor=COLOR["header_bg"])
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    for i, (name, width) in enumerate(SUMMARY_COLS, 1):
        _hcell(ws, 2, i, name, bg=COLOR["ti_blue"])
        ws.column_dimensions[get_column_letter(i)].width = width

    r = 3
    for res in results:
        eol  = res.eol_part
        best = res.ti_alternatives[0] if res.ti_alternatives else None
        bg   = _row_bg(best.match_type if best else "")

        def d(col, val, **kw):
            return _dcell(ws, r, col, val, bg=bg, **kw)

        d(1, eol.part_number, bold=True)
        d(2, eol.supplier)
        d(3, eol.description, wrap=True)
        d(4, _fmt_date(eol.eol_date))
        d(5, _fmt_date(eol.last_buy_date))
        d(6, eol.replacement_pn)

        if best:
            d(7,  best.ti_part_number, bold=True)
            d(8,  best.ti_description, wrap=True)
            d(9,  best.match_type, bold=True,
              color=COLOR["green_txt"] if "Drop" in best.match_type else "000000")
            d(10, best.lifecycle_status)
            p = best.params or {}
            d(11, p.get("Vin_min", "") or p.get("Vin(min)", ""))
            d(12, p.get("Vin_max", "") or p.get("Vin(max)", ""))
            d(13, p.get("Vout_min", "") or p.get("Vout(min)", ""))
            d(14, p.get("Vout_max", "") or p.get("Vout(max)", ""))
            d(15, p.get("Iout", "") or p.get("Iout(max)", ""))
            d(16, p.get("Package", ""))
            _dcell(ws, r, 17, best.ti_part_number, bg=bg, link=best.ti_product_url)
        else:
            for col in range(7, 17):
                d(col, "No TI match found" if col == 7 else "")
            d(17, "")

        src_label, src_url = _source(eol)
        _dcell(ws, r, 18, src_label, bg=bg, link=src_url if src_url else None)
        ws.row_dimensions[r].height = 20
        r += 1

    # Sheet 2 — Full Detail
    ws2 = wb.create_sheet("Full Detail")
    ws2.freeze_panes = "A3"
    d2_cols = [("EOL Part",18),("Supplier",16),("TI Alternative",20),
               ("Match Type",20),("Parameter",24),("TI Value",18),
               ("EOL Source / Reference",50)]
    ws2.merge_cells(f"A1:{get_column_letter(len(d2_cols))}1")
    t2 = ws2.cell(row=1, column=1, value="Full Cross-Reference Detail")
    t2.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    t2.fill = PatternFill("solid", fgColor=COLOR["header_bg"])
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 32
    for i, (name, width) in enumerate(d2_cols, 1):
        _hcell(ws2, 2, i, name, bg=COLOR["ti_blue"])
        ws2.column_dimensions[get_column_letter(i)].width = width

    r2 = 3
    for res in results:
        eol = res.eol_part
        src_label, src_url = _source(eol)
        for alt in res.ti_alternatives:
            bg = _row_bg(alt.match_type)
            params = list((alt.params or {}).items()) or [("(no parameter data)", "")]
            for j, (pk, pv) in enumerate(params):
                row_bg = bg if j == 0 else (COLOR["row_white"] if r2 % 2 else COLOR["row_alt"])
                _dcell(ws2, r2, 1, eol.part_number if j == 0 else "", bg=row_bg, bold=(j == 0))
                _dcell(ws2, r2, 2, eol.supplier    if j == 0 else "", bg=row_bg)
                _dcell(ws2, r2, 3, alt.ti_part_number if j == 0 else "", bg=row_bg, bold=(j == 0))
                _dcell(ws2, r2, 4, alt.match_type  if j == 0 else "", bg=row_bg, bold=(j == 0),
                       color=COLOR["green_txt"] if "Drop" in alt.match_type else "000000")
                _dcell(ws2, r2, 5, pk, bg=row_bg)
                _dcell(ws2, r2, 6, pv, bg=row_bg)
                _dcell(ws2, r2, 7, src_label if j == 0 else "", bg=row_bg,
                       link=src_url if (j == 0 and src_url) else None)
                ws2.row_dimensions[r2].height = 18
                r2 += 1

    # Sheet 3 — No Match
    no_match = [res for res in results if not res.ti_alternatives]
    ws3 = wb.create_sheet(f"No TI Match ({len(no_match)})")
    nm_cols = [("EOL Part",22),("Supplier",20),("Description",40),
               ("EOL Date",14),("Last Buy Date",14),
               ("Supplier Replacement",22),("EOL Source / Reference",50)]
    ws3.merge_cells(f"A1:{get_column_letter(len(nm_cols))}1")
    t3 = ws3.cell(row=1, column=1, value=f"No TI Match Found — {len(no_match)} parts")
    t3.font = Font(bold=True, size=12, color="FFFFFF", name="Arial")
    t3.fill = PatternFill("solid", fgColor="843C0C")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 22
    ws3.row_dimensions[2].height = 32
    for i, (name, width) in enumerate(nm_cols, 1):
        _hcell(ws3, 2, i, name, bg="C55A11")
        ws3.column_dimensions[get_column_letter(i)].width = width

    for r3, res in enumerate(no_match, 3):
        eol = res.eol_part
        bg = COLOR["row_white"] if r3 % 2 else COLOR["row_alt"]
        for col, val in enumerate([eol.part_number, eol.supplier, eol.description,
                                    _fmt_date(eol.eol_date), _fmt_date(eol.last_buy_date),
                                    eol.replacement_pn], 1):
            _dcell(ws3, r3, col, val, bg=bg, bold=(col == 1), wrap=(col == 3))
        src_label, src_url = _source(eol)
        _dcell(ws3, r3, 7, src_label, bg=bg, link=src_url if src_url else None)
        ws3.row_dimensions[r3].height = 20

    wb.save(output_path)
    log.info(f"Report saved: {output_path}")
    return output_path
